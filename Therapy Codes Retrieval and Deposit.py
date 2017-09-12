'''
Postcondition 1: Program takes all relevant data (facility name, patient name, visits, minutes, status codes) from all therapy records in chosen directory and places it in the Therapy Dashboard.xlsm
in a newly created sheet named with the facility, therapy discipline, and month
'''
from tkinter import *

import os
import subprocess
import openpyxl
import tkinter
from tkinter import filedialog
import time

# excel_document = openpyxl.load_workbook(filename)

def main():
    Tk().withdraw # we don't want a full GUI, so keep the root window from appearing

    filedirectory = filedialog.askdirectory(title="Open Folder",
     initialdir=('G:\\Therapy Charting Grids\\'))

    therapy_dashboard = openpyxl.load_workbook('G:\Therapy Record Interfaces\Directors of Therapy\Centre Avenue\Therapy Code Dashboard.xlsm', read_only=False, keep_vba=True)
    therapy_sheet = therapy_dashboard.sheetnames
    therapy_sheet = therapy_sheet[0]
    therapy_sheet = therapy_dashboard[therapy_sheet]
    therapy_sheet.sheet_state = 'visible'

    os.chdir(filedirectory)

    discipline = "Null"
    
    # Subgoal 1: Here we are defining month/discipline/facility variables from the chosen file path stored in the variable filedirectory
    if " January " in filedirectory:
        month = "January"
    elif " February " in filedirectory:
        month = "February"
    elif " March " in filedirectory:
        month = "March"
    elif " April " in filedirectory:
        month = "April"
    elif " May " in filedirectory:
        month = "May"
    elif " June " in filedirectory:
        month = "June"
    elif " July " in filedirectory:
        month = "July"
    elif " August " in filedirectory:
        month = "August"
    elif " September " in filedirectory:
        month = "September"
    elif " October " in filedirectory:
        month = "October"
    elif " November " in filedirectory:
        month = "November"
    elif " December " in filedirectory:
        month = "December"
    if (r"/OT") in filedirectory:
        discipline = "OT"
    elif (r"/PT") in filedirectory:
        discipline = "PT"
    elif (r"/ST") in filedirectory:
        discipline = "ST"
    if "Centre Avenue" in filedirectory:
        facility = "Centre Avenue"
    elif "Columbine Commons" in filedirectory:
        facility = "Columbine Commons"
    elif (r"North Shore") in filedirectory:
        facility = "North Shore"
    elif "Lemay Avenue" in filedirectory:
        facility = "Lemay Avenue"
    elif "Columbine West" in filedirectory:
        facility = "Columbine West"
    # Subgoal 2: Create new worksheet from template(sheet1) rename according to month/discipline/facility variables and use worksheet to to hold all values retrieved from therapy records
    therapy_sheet_template = therapy_sheet
    therapy_dashboard.copy_worksheet(therapy_sheet)
    therapy_sheet_template.state = 'hidden'
    therapy_sheet = therapy_dashboard.sheetnames
    therapy_sheet = therapy_sheet[-1]
    therapy_sheet = therapy_dashboard[therapy_sheet]
    therapy_sheet.title = month + " " + facility
    therapy_sheet = month + " " + facility
    # Postcondition 1: All relevant data is retrieved from therapy records and placed in worksheet
    for folders, directory, files in os.walk(filedirectory):
        for filename in files:
            if ".xlsm" in filename and "~" not in filename:
                print("Processing ", filename)
                filename_directory = os.path.join(folders, filename)
                list_G0283 = []
                list_97024 = []
                list_97110 = []
                list_97032 = []
                list_97035 = []
                therapy_code_lists = [ list_G0283, list_97024, list_97110, list_97032, list_97035 ]
                excel_document = openpyxl.load_workbook(filename=filename_directory, read_only=True, data_only=True)
                sheet = excel_document.get_sheet_names()[0]
                patient_first_name = excel_document[sheet]["O3"].value
                patient_last_name = excel_document[sheet]["H3"].value
                patient_name = "{}, {}".format(patient_last_name, patient_first_name)
                named_ranges = excel_document.get_named_ranges()
                if " OT" in filename:
                    discipline = "OT"
                elif " PT" in filename:
                    discipline = "PT"
                elif " ST" in filename:
                    discipline = "ST"
                for named_range in named_ranges:
                    address_split = named_range.attr_text.split("!")
                    address = address_split[1]
                    if named_range.name == "TotalMinutes":
                        total_minutes = excel_document[sheet][address].value
                    if named_range.name == "PreviousMonthVisits":
                        previous_month_visits = excel_document[sheet][address].value
                    if named_range.name == "CurrentMonthVisits":
                        current_month_visits = excel_document[sheet][address].value
                cell_range = excel_document[sheet]['A8':'A32']
                for cell in cell_range:
                    try:
                        if 'G0283' in cell[0].value:
                            row_G0283 = cell[0].row
                        if '97024' in cell[0].value:
                            row_97024 = cell[0].row
                        if '97110' in cell[0].value:
                            row_97110 = cell[0].row
                        if '97032' in cell[0].value:
                            row_97032 = cell[0].row
                        if '97035' in cell[0].value:
                            row_97035 = cell[0].row
                    except TypeError:
                        pass
                row_lists = [ row_G0283, row_97024, row_97110, row_97032, row_97035 ]
                for num in range(2, 33):
                    for code_list in therapy_code_lists:
                        code_list.append(excel_document[sheet].cell(row=row_lists[therapy_code_lists.index(code_list)], column=num).value)
                #for code_list in therapy_code_lists:
                    #for item in code_list:
                        #pass

                try:
                    total_visits = previous_month_visits + current_month_visits
                except TypeError:
                    total_visits = current_month_visits

                for num in range(1, 5000):
                    if therapy_dashboard[therapy_sheet].cell(row=num, column=6).value is None:
                        therapy_dashboard[therapy_sheet].cell(row=num, column=3).value = filename
                        therapy_dashboard[therapy_sheet].cell(row=num, column=4).value = patient_name
                        therapy_dashboard[therapy_sheet].cell(row=num, column=5).value = discipline
                        therapy_dashboard[therapy_sheet].cell(row=num, column=43).value = total_minutes
                        therapy_dashboard[therapy_sheet].cell(row=num, column=6).value = "G0283 E-stim unattended"
                        therapy_dashboard[therapy_sheet].cell(row=num+1, column=6).value = "97024 SWD Supervised"
                        therapy_dashboard[therapy_sheet].cell(row=num+2, column=6).value = "97110:Ther-ex ROM/Strength"
                        therapy_dashboard[therapy_sheet].cell(row=num+3, column=6).value = "97032 Estim Attended"
                        therapy_dashboard[therapy_sheet].cell(row=num+4, column=6).value = "97035: Ultrasound"
                        therapy_dashboard[therapy_sheet].cell(row=num+5, column=6).value = ""
                        for num_2 in range(0, 31):
                            therapy_dashboard[therapy_sheet].cell(row=num, column=num_2+7).value = therapy_code_lists[0][num_2]
                            therapy_dashboard[therapy_sheet].cell(row=num+1, column=num_2+7).value = therapy_code_lists[1][num_2]
                            therapy_dashboard[therapy_sheet].cell(row=num+2, column=num_2+7).value = therapy_code_lists[2][num_2]
                            therapy_dashboard[therapy_sheet].cell(row=num+3, column=num_2+7).value = therapy_code_lists[3][num_2]
                            therapy_dashboard[therapy_sheet].cell(row=num+4, column=num_2+7).value = therapy_code_lists[4][num_2]
                        break

            #a = input("Continue? Y/n: ")
            #print(a)
            #if a == "Y":
                 #pass
            #else:
                #break

    home_dir = os.path.normpath("G:/Therapy Record Interfaces/Directors of Therapy/Centre Avenue/")
    os.chdir(home_dir)
    therapy_sheet = therapy_dashboard.sheetnames
    therapy_sheet = therapy_sheet[0]
    therapy_sheet_template = therapy_dashboard[therapy_sheet]
    therapy_sheet_template.sheet_state = 'hidden'
    therapy_dashboard.save('Therapy Code Dashboard.xlsm')

    os.startfile('G:\Therapy Record Interfaces\Directors of Therapy\Centre Avenue\Therapy Code Dashboard.xlsm')

main()

