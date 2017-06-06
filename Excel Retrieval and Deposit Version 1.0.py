try:
    # for Python2
    from Tkinter import *   ## notice capitalized T in Tkinter 
except ImportError:
    # for Python3
    from tkinter import *   ## notice lowercase 't' in tkinter here

import os
import subprocess
import openpyxl
import tkinter
from tkinter import filedialog

#OPTIONS = [
#    ]

#root = Tk()

#var1 = tkinter.StringVar()
#drop = tkinter.OptionMenu(root,var1,*OPTIONS)
#drop.grid()

Tk().withdraw # we don't want a full GUI, so keep the root window from appearing

filedirectory = filedialog.askdirectory(title="Open Folder",
 initialdir=('G:\\Therapy Charting Grids\\'))

therapy_dashboard = openpyxl.load_workbook('G:\Therapy Record Interfaces\Front Range Therapy\Dashboard\Therapy Dashboard.xlsm', keep_vba=True)
therapy_sheet = therapy_dashboard.get_sheet_names()[0]

os.chdir(filedirectory)

# excel_document = openpyxl.load_workbook(filename)

def main():
    for filename in os.listdir(filedirectory):
        if ".xlsm" in filename and "~" not in filename:
            print("Processing", filename)
            excel_document = openpyxl.load_workbook(filename=filename, read_only=True, data_only=True)
            sheet = excel_document.get_sheet_names()[0]
            patient_first_name = excel_document[sheet]["O3"].value
            patient_last_name = excel_document[sheet]["H3"].value
            patient_name = "{}, {}".format(patient_last_name, patient_first_name)
            named_ranges = excel_document.get_named_ranges()
            discipline = "Null"
            visits_dict = {}
            if " OT" in filename:
                discipline = "OT"
            elif " PT" in filename:
                discipline = "PT"
            elif " ST" in filename:
                discipline = "ST"
            facility_name = excel_document[sheet]["W2"].value
            for named_range in named_ranges:
                address_split = named_range.attr_text.split("!")
                address = address_split[1]
                if named_range.name == "TotalMinutes":
                    total_minutes = excel_document[sheet][address].value
                if named_range.name == "PreviousMonthVisits":
                    previous_month_visits = excel_document[sheet][address].value
                if named_range.name == "CurrentMonthVisits":
                    current_month_visits = excel_document[sheet][address].value

                cell_dict = {"B": 6, "C": 7, "D": 8, "E": 9, "F": 10, "G": 11, "H": 12, "I": 13, "J": 14, "K": 15, "L": 16, "M": 17, "N": 18,
                             "O": 19, "P": 20, "Q": 21, "R": 22,"S": 23, "T": 24, "U": 25, "V": 26, "W": 27, "X": 28, "Y": 29, "Z": 30, "AA": 31,
                             "AB": 32, "AC": 33, "AD": 34, "AE": 35, "AF": 36}
                for cell in cell_dict.keys():
                    if named_range.name == "{}{}{}".format("Initials", cell, "1"):
                        if excel_document[sheet][address].value is None:
                            pass
                        else:
                            visits_dict["{}{}{}".format("Initials", cell, "1")] = "X"
                            break
                    if named_range.name == "{}{}".format("TreatmentMinutes", cell):
                        if excel_document[sheet][address].value == "R":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "R"
                        elif excel_document[sheet][address].value == "A":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "A"
                        elif excel_document[sheet][address].value == "B":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "B"
                        elif excel_document[sheet][address].value == "C":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "C"
                        elif excel_document[sheet][address].value == "D":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "D"
            try:
                total_visits = previous_month_visits + current_month_visits
            except TypeError:
                total_visits = current_month_visits
            for num in range(1, 1000):
                    if therapy_dashboard[therapy_sheet].cell(row=num, column=1).value is None:
                        therapy_dashboard[therapy_sheet].cell(row=num, column=1).value = facility_name
                        therapy_dashboard[therapy_sheet].cell(row=num, column=2).value = patient_name
                        therapy_dashboard[therapy_sheet].cell(row=num, column=3).value = discipline
                        therapy_dashboard[therapy_sheet].cell(row=num, column=4).value = total_visits
                        therapy_dashboard[therapy_sheet].cell(row=num, column=5).value = total_minutes
                        for cell, x in visits_dict.items():
                            for cell_2, num_2 in cell_dict.items():
                                if cell == "{}{}{}".format("Initials", cell_2, "1"):
                                    therapy_dashboard[therapy_sheet].cell(row=num, column=num_2).value = x
                        for cell, x in visits_dict.items():
                            for cell_2, num_2 in cell_dict.items():
                                if cell == "{}{}".format("TreatmentMinutes", cell_2):
                                    therapy_dashboard[therapy_sheet].cell(row=num, column=num_2).value = x
                        break

    home_dir = os.path.normpath("G:/Therapy Record Interfaces/Front Range Therapy/Dashboard/")
    os.chdir(home_dir)
    therapy_dashboard.save('Therapy Dashboard.xlsm')

    os.startfile('G:\Therapy Record Interfaces\Front Range Therapy\Dashboard\Therapy Dashboard.xlsm')
    #raise KeyError(filename, named_range, 'not found')

main()

#print(excel_document.get_sheet_names())

#sheet = excel_document.get_sheet_by_name('Sheet1')
#print(sheet.cell(row = 5, column = 2).value)

#multiple_cells = sheet['A1':'B3']
#for row in multiple_cells:
#    for cell in row:
#        print(cell.value)

#all_rows = sheet.rows
#print(all_rows)

#all_columns = sheet.columns
#print(all_columns)
